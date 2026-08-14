from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ExcelWorkbookLoader:
    """Carga y valida la matriz de criterios desde un archivo Excel."""

    def __init__(self, file_path: str | Path | None = None):
        self.file_path = Path(file_path) if file_path is not None else None

    def load(self, file_path: str | Path) -> dict[str, Any]:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"No se encontró el archivo Excel: {path}")

        workbook = load_workbook(path, data_only=True)
        if not workbook.sheetnames:
            raise ValueError("El archivo Excel no contiene hojas.")
        return {sheet.title: sheet for sheet in workbook.worksheets}

    def extract_materials_and_criteria(self, workbook_data: dict[str, Any]) -> tuple[list[str], list[str], dict[str, list[float]]]:
        if not workbook_data:
            raise ValueError("No hay información en el archivo Excel.")

        material_sheet = self._find_sheet(workbook_data, ["Materiales", "materials", "MATERIALES"])
        criteria_sheet = self._find_sheet(workbook_data, ["Criterios", "criteria", "CRITERIOS"])

        materials = self._read_named_list(material_sheet, 1)
        criteria = self._read_named_list(criteria_sheet, 0)

        matrix: dict[str, list[float]] = {}
        for column_index, criterion in enumerate(criteria):
            values = []
            for row_index in range(1, material_sheet.max_row + 1):
                value = material_sheet.cell(row=row_index, column=column_index + 2).value
                if value is None:
                    continue
                values.append(float(value))
            matrix[criterion] = values

        return materials, criteria, matrix

    def _find_sheet(self, workbook_data: dict[str, Any], aliases: list[str]):
        for alias in aliases:
            for sheet_name, sheet in workbook_data.items():
                if sheet_name.strip().lower() == alias.lower():
                    return sheet
        for sheet_name, sheet in workbook_data.items():
            lower = sheet_name.lower()
            if "material" in lower or "criterio" in lower:
                return sheet
        raise ValueError(f"No se encontró ninguna hoja con nombre esperado: {aliases}")

    def _read_named_list(self, sheet: Any, column_index: int) -> list[str]:
        names: list[str] = []
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            value = row[column_index] if column_index < len(row) else None
            if value is None:
                continue
            text = str(value).strip()
            if not text or text.lower() in {"material", "criterio", "total", "#", "datos"}:
                continue
            names.append(text)
        return names
