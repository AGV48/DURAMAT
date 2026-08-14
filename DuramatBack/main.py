from __future__ import annotations

from io import BytesIO

import openpyxl
from fastapi import FastAPI, File, Form, UploadFile
from fastapi.exceptions import RequestValidationError
from fastapi import Request
import logging
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse

from app.config import settings
from app.models.schemas import ClimateInput, EvaluationResponse, MaterialEvaluation
from app.services.decision_model import DecisionModelEngine

app = FastAPI(title=settings.app_name, version=settings.app_version, debug=settings.debug)

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.allowed_origins or ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(ValueError)
async def value_error_handler(_, exc: ValueError):
    return JSONResponse(status_code=400, content={"detail": str(exc)})


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    logging.getLogger("uvicorn.error").error("Request validation failed for %s: %s", request.url.path, exc)
    return JSONResponse(status_code=400, content={"detail": "Request validation error", "errors": exc.errors()})


@app.get("/health")
def health_check() -> dict[str, str]:
    return {"status": "ok", "app": settings.app_name}


@app.post("/api/evaluate", response_model=EvaluationResponse)
async def evaluate_excel(
    file: UploadFile = File(...),
    temperature_c: float = Form(24.0),
    relative_humidity: float = Form(75.0),
    co2_ppm: float = Form(420.0),
) -> EvaluationResponse:
    # Log incoming basic information to help debug 400 responses (avoid uvicorn access logger formatting)
    try:
        print(f"Incoming evaluate request: filename={getattr(file, 'filename', None)}, temperature_c={temperature_c}, relative_humidity={relative_humidity}, co2_ppm={co2_ppm}")
    except Exception:
        pass
    if not file.filename:
        raise ValueError("Debe adjuntar un archivo Excel.")

    content = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)

    material_names: list[str] = []
    criteria: list[str] = []
    matrix: dict[str, list[float]] = {}

    # Prefer a sheet named like 'Matriz Datos Entrada' (common in provided workbook)
    matrix_sheet = None
    for s in workbook.worksheets:
        title = s.title.lower()
        if "matriz" in title and "datos" in title:
            matrix_sheet = s
            break

    if matrix_sheet is not None:
        # locate header row that contains 'Material'
        header_row_index = None
        header = None
        for i, row in enumerate(matrix_sheet.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
            if any(isinstance(c, str) and c.strip().lower() in {"material", "materiales"} for c in row if c is not None):
                header_row_index = i
                header = row
                break

        if header_row_index is not None and header is not None:
            try:
                material_col_index = next(idx for idx, c in enumerate(header) if isinstance(c, str) and c.strip().lower() in {"material", "materiales"})
            except StopIteration:
                material_col_index = 1

            # extract material names from rows below header
            for row in matrix_sheet.iter_rows(min_row=header_row_index + 1, max_row=matrix_sheet.max_row, values_only=True):
                v = row[material_col_index] if material_col_index < len(row) else None
                if v is None:
                    continue
                s = str(v).strip()
                if s and s.lower() not in {"material", "materiales", "total", "#", "datos"}:
                    material_names.append(s)

            # extract criteria as columns after the material column
            for col_idx in range(material_col_index + 1, len(header)):
                name = header[col_idx]
                if not name:
                    continue
                name_str = str(name).strip()
                values: list[float] = []
                for row in matrix_sheet.iter_rows(min_row=header_row_index + 1, max_row=header_row_index + len(material_names), values_only=True):
                    v = row[col_idx] if col_idx < len(row) else None
                    try:
                        values.append(float(v))
                    except Exception:
                        values.append(None)
                if len(values) == len(material_names) and all(v is not None for v in values):
                    criteria.append(name_str)
                    matrix[name_str] = values
        else:
            # Heuristic fallback: find the column that looks like material names (most string values)
            try:
                col_string_counts = []
                for col_idx in range(0, matrix_sheet.max_column):
                    count = 0
                    for r in matrix_sheet.iter_rows(min_row=1, max_row=min(60, matrix_sheet.max_row), values_only=True):
                        v = r[col_idx] if col_idx < len(r) else None
                        if v is None:
                            continue
                        # prefer textual cells containing letters (avoid numeric markers like '1','2' or short tokens)
                        if isinstance(v, str):
                            s = v.strip()
                            if s and any(ch.isalpha() for ch in s) and not any(k in s.lower() for k in ("materiales", "total", "datos")):
                                count += 1
                    col_string_counts.append((col_idx, count))
                col_string_counts.sort(key=lambda x: x[1], reverse=True)
                if col_string_counts and col_string_counts[0][1] >= 2:
                    material_col_index = col_string_counts[0][0]
                    # find first row with a likely material name
                    start_row = None
                    for i, r in enumerate(matrix_sheet.iter_rows(min_row=1, max_row=matrix_sheet.max_row, values_only=True), start=1):
                        v = r[material_col_index] if material_col_index < len(r) else None
                        if v is None:
                            continue
                        s = str(v).strip()
                        if s and not any(k in s.lower() for k in ("material", "materiales", "total", "datos", "#")):
                            start_row = i
                            break
                    if start_row:
                        # treat the row above as header if possible
                        header_row_index = max(1, start_row - 1)
                        header = [c for c in next(matrix_sheet.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))]
                        # collect material names from start_row onward
                        material_names = []
                        for row in matrix_sheet.iter_rows(min_row=start_row, max_row=matrix_sheet.max_row, values_only=True):
                            v = row[material_col_index] if material_col_index < len(row) else None
                            if v is None:
                                continue
                            s = str(v).strip()
                            if s and not any(k in s.lower() for k in ("material", "materiales", "total", "datos", "#")):
                                material_names.append(s)
                        # attempt to extract criteria as columns using the header we found
                        for col_idx in range(0, len(header)):
                            if col_idx == material_col_index:
                                continue
                            name = header[col_idx] if col_idx < len(header) else None
                            if not name:
                                continue
                            name_str = str(name).strip()
                            values = []
                            for row in matrix_sheet.iter_rows(min_row=start_row, max_row=start_row + len(material_names) - 1, values_only=True):
                                v = row[col_idx] if col_idx < len(row) else None
                                try:
                                    values.append(float(v))
                                except Exception:
                                    values.append(None)
                            if len(values) == len(material_names) and all(v is not None for v in values):
                                criteria.append(name_str)
                                matrix[name_str] = values
            except Exception:
                pass

    # Fallbacks: previous heuristics (sheet name contains 'Material' or 'Criterio')
    if not material_names:
        for sheet in workbook.worksheets:
            sheet_name_lower = sheet.title.lower()
            if "material" in sheet_name_lower or "materiales" in sheet_name_lower:
                material_names = [
                    str(value).strip()
                    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True)
                    for value in [row[1] if len(row) > 1 else row[0] if len(row) > 0 else ""]
                    if str(value).strip()
                    and str(value).strip().lower() not in {"material", "materiales", "total", "#", "datos"}
                ]
                if material_names:
                    break

    if not criteria:
        # row-based criteria parsing: rows where first cell is name and following numeric values match material count
        for sheet in workbook.worksheets:
            parsed = [
                row
                for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True)
                if row and row[0] is not None and str(row[0]).strip()
            ]
            parsed_criteria: list[tuple[str, list[float]]] = []
            for row in parsed:
                name = str(row[0]).strip()
                if not name or name.lower() in {"criterio", "criterios", "total", "#", "datos"}:
                    continue
                values: list[float] = []
                for v in row[1:]:
                    if v is None:
                        continue
                    try:
                        values.append(float(v))
                    except Exception:
                        continue
                if values and ((material_names and len(values) == len(material_names)) or (not material_names and len(values) > 1)):
                    parsed_criteria.append((name, values))

            if parsed_criteria:
                criteria = [name for name, _ in parsed_criteria]
                for name, values in parsed_criteria:
                    matrix[name] = values
                break

    if not criteria and material_names:
        # column-based criteria parsing: prefer sheets that look like 'Matriz' or 'Datos'
        preferred = [s for s in workbook.worksheets if ("matriz" in s.title.lower() and "datos" in s.title.lower())]
        if not preferred:
            preferred = [s for s in workbook.worksheets if any(k in s.title.lower() for k in ("matriz", "datos"))]
        others = [s for s in workbook.worksheets if s not in preferred]
        search_order = preferred + others
        for sheet in search_order:
            header_row_index = None
            header = None
            for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
                if any(isinstance(c, str) and c.strip().lower() in {'material', 'materiales'} for c in row if c is not None):
                    header_row_index = i
                    header = row
                    break
            if header_row_index is None:
                continue

            try:
                material_col_index = next(idx for idx, c in enumerate(header) if isinstance(c, str) and c.strip().lower() in {'material', 'materiales'})
            except StopIteration:
                continue

            found = False
            for col_idx in range(material_col_index + 1, len(header)):
                name = header[col_idx]
                if name is None:
                    continue
                name_str = str(name).strip()
                if not name_str:
                    continue
                values = []
                for row in sheet.iter_rows(min_row=header_row_index + 1, max_row=header_row_index + len(material_names), values_only=True):
                    v = row[col_idx] if col_idx < len(row) else None
                    if v is None:
                        values.append(None)
                    else:
                        try:
                            values.append(float(v))
                        except Exception:
                            values.append(None)
                if len(values) == len(material_names) and all(v is not None for v in values):
                    matrix[name_str] = values
                    criteria.append(name_str)
                    found = True

            if found:
                break

    # Debug dump immediately after parsing to inspect shapes before validation
    try:
        print(f"PARSE_DEBUG material_names ({len(material_names)}): {material_names}")
        print(f"PARSE_DEBUG criteria ({len(criteria)}): {criteria}")
        for k, v in matrix.items():
            print(f"PARSE_DEBUG matrix[{k}] len={len(v)} sample={v[:8]}")
    except Exception:
        pass

    if not material_names:
        raise ValueError("No se encontraron materiales en la hoja del Excel. Verifica que el archivo tenga una hoja con nombre que incluya 'Material' o 'Materiales'.")

    if not criteria:
        raise ValueError("No se encontraron criterios en la hoja del Excel. Verifica que la hoja con criterios tenga valores válidos en la primera columna o como encabezados de columna.")

    # Map descriptive column headers to internal criterion keys expected by DecisionModelEngine
    def map_criterion_name(name: str) -> str | None:
        s = name.strip().lower()
        if any(k in s for k in ("co2", "gwp", "emisiones")):
            return "co2"
        if any(k in s for k in ("energ", "energia")):
            return "energy"
        if any(k in s for k in ("técnic", "tecnic", "desempe")):
            return "technical_performance"
        if any(k in s for k in ("costo", "lcc")):
            return "lcc_cost"
        if any(k in s for k in ("salud", "ecosist")):
            return "health_ecosystems"
        if any(k in s for k in ("mencion", "cantidad")):
            return "cantidad_menciones"
        return None

    mapped_matrix: dict[str, list[float]] = {}
    for original_name, values in matrix.items():
        mapped = map_criterion_name(original_name)
        if mapped:
            # avoid overwriting if duplicates
            if mapped in mapped_matrix:
                # append suffix to keep original
                mapped_matrix[f"{mapped}_alt"] = values
            else:
                mapped_matrix[mapped] = values
        else:
            # preserve unknown criteria under original name
            mapped_matrix[original_name] = values

    climate = ClimateInput(temperature_c=temperature_c, relative_humidity=relative_humidity, co2_ppm=co2_ppm)
    engine = DecisionModelEngine()
    # Debug: log parsed matrices before validation to diagnose 400 responses
    try:
        print(f"DEBUG parsed material_names ({len(material_names)}): {material_names}")
        print(f"DEBUG parsed criteria ({len(criteria)}): {criteria}")
        for k, v in matrix.items():
            print(f"DEBUG matrix[{k}] len={len(v)} sample={v[:8]}")
    except Exception:
        import traceback

        print("Failed to log debug parsing info:\n", traceback.format_exc())
    engine.validate_inputs(mapped_matrix, material_names)
    raw_ranking, tree = engine.calculate_scores(material_names, mapped_matrix, climate.model_dump())
    engine.print_decision_tree(tree)

    ranking = [
        MaterialEvaluation(
            rank=item["rank"],
            material=item["material"],
            score=float(item["score"]),
            life_years=float(item["life_years"]),
            annualized_co2=float(item["annualized_co2"]),
            annualized_energy=float(item["annualized_energy"]),
            technical_performance=float(item["technical_performance"]),
            co2=float(item["co2"]),
            energy=float(item["energy"]),
            lcc_cost=float(item["lcc_cost"]),
            health_ecosystems=float(item["health_ecosystems"]),
            contribution={k: float(v) for k, v in item["contribution"].items()},
        )
        for item in raw_ranking
    ]

    top_material = ranking[0].material if ranking else None
    score_gap = 0.0
    if len(ranking) > 1:
        score_gap = ((ranking[0].score - ranking[1].score) / max(ranking[0].score, 1e-9)) * 100

    return EvaluationResponse(
        status="success",
        message="Evaluación completada exitosamente.",
        climate=climate,
        ranking=ranking,
        top_material=top_material,
        score_gap_percent=score_gap,
    )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
