from openpyxl import load_workbook

wb = load_workbook('DuramatFront/MCDM_Materiales.xlsx', read_only=True)
ms = None
for s in wb.worksheets:
    if 'matriz' in s.title.lower() and 'datos' in s.title.lower():
        ms = s
        break
if ms is None:
    print('no matrix sheet')
    raise SystemExit(0)

# find header row containing 'Material'
header_row_index = None
header = None
for i, row in enumerate(ms.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
    if any(isinstance(c, str) and 'material' in c.lower() for c in row if c is not None):
        header_row_index = i
        header = row
        break
print('detected header_row_index:', header_row_index)
print('header row:', header)

# material_col_index
if header is not None:
    try:
        material_col_index = next(idx for idx, c in enumerate(header) if isinstance(c, str) and 'material' in c.lower())
    except StopIteration:
        material_col_index = 1
    print('material_col_index:', material_col_index)

    # show next rows values at that column
    print('next rows at material_col:')
    for i, row in enumerate(ms.iter_rows(min_row=header_row_index+1, max_row=header_row_index+12, values_only=True), start=header_row_index+1):
        v = row[material_col_index] if material_col_index < len(row) else None
        print(i, v)

# fallback heuristic counts
col_string_counts = []
for col_idx in range(0, ms.max_column):
    count = 0
    for r in ms.iter_rows(min_row=1, max_row=min(60, ms.max_row), values_only=True):
        v = r[col_idx] if col_idx < len(r) else None
        if v is None:
            continue
        if isinstance(v, str):
            s = v.strip()
            if s and any(ch.isalpha() for ch in s) and not any(k in s.lower() for k in ("materiales", "total", "datos")):
                count += 1
    col_string_counts.append((col_idx, count))

col_string_counts.sort(key=lambda x: x[1], reverse=True)
print('col_string_counts:', col_string_counts)

# show column samples
for col_idx, cnt in col_string_counts[:5]:
    print('col', col_idx, 'sample:')
    for i, row in enumerate(ms.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
        v = row[col_idx] if col_idx < len(row) else None
        print(i, v)
    print()
