from openpyxl import load_workbook

wb = load_workbook('DuramatFront/MCDM_Materiales.xlsx', read_only=True)
for s in wb.worksheets:
    title = s.title
    print('--- SHEET:', title)
    if 'matriz' in title.lower() and 'datos' in title.lower():
        print('Found matrix sheet:', title)
        for i, row in enumerate(s.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
            vals = [str(c) if c is not None else '' for c in row]
            print(f'{i:02d}:', vals)
    else:
        # print first 3 rows of other sheets
        for i, row in enumerate(s.iter_rows(min_row=1, max_row=3, values_only=True), start=1):
            vals = [str(c) if c is not None else '' for c in row]
            print(f'  {i:02d}:', vals)
    print()
