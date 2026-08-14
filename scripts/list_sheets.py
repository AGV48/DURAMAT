from openpyxl import load_workbook

wb = load_workbook('DuramatFront/MCDM_Materiales.xlsx', read_only=True)
print('SHEETS:')
for name in wb.sheetnames:
	print('-', name)

for name in wb.sheetnames:
	sheet = wb[name]
	print('\n---', name, '---')
	for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
		print(i, row)
