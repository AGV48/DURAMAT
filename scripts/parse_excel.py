from openpyxl import load_workbook
from io import BytesIO

wb = load_workbook('DuramatFront/MCDM_Materiales.xlsx', read_only=True)

material_names = []
criteria = []
matrix = {}

for sheet in wb.worksheets:
    sheet_name_lower = sheet.title.lower()

    if "material" in sheet_name_lower or "materiales" in sheet_name_lower:
        material_names = [
            str(value).strip()
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True)
            for value in [row[1] if len(row) > 1 else row[0] if len(row) > 0 else ""]
            if str(value).strip()
        matrix_sheet = None
        for s in wb.worksheets:
            title = s.title.lower()
            if "matriz" in title and "datos" in title:
                matrix_sheet = s
                break
        if matrix_sheet is not None:
            header_row_index = None
            header = None
            for i, row in enumerate(matrix_sheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
                if any(isinstance(c, str) and 'material' in c.lower() for c in row if c is not None):
                    header_row_index = i
                    header = row
                    break
            if header_row_index is not None and header is not None:
                try:
                    material_col_index = next(idx for idx, c in enumerate(header) if isinstance(c, str) and 'material' in c.lower())
                except StopIteration:
                    material_col_index = 1

                material_names = []
                for row in matrix_sheet.iter_rows(min_row=header_row_index + 1, max_row=matrix_sheet.max_row, values_only=True):
                    v = row[material_col_index] if material_col_index < len(row) else None
                    if v is None:
                        continue
                    s = str(v).strip()
                    if s and s.lower() not in {"material", "materiales", "total", "#", "datos"}:
                        material_names.append(s)

                for col_idx in range(material_col_index + 1, len(header)):
                    name = header[col_idx]
                    if not name:
                        continue
                    name_str = str(name).strip()
                    values = []
                    for row in matrix_sheet.iter_rows(min_row=header_row_index + 1, max_row=matrix_sheet.max_row, values_only=True):
                        v = row[col_idx] if col_idx < len(row) else None
                        if v is None:
                            continue
                        try:
                            values.append(float(v))
                        except Exception:
                            continue
                    if values and len(values) == len(material_names):
                        criteria.append(name_str)
                        matrix[name_str] = values

            and str(value).strip().lower() not in {"material", "materiales", "total", "#", "datos"}
        ]

    if "criterio" in sheet_name_lower or "criterios" in sheet_name_lower:
        criteria_rows = [
            row
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True)
            if row and row[0] is not None and str(row[0]).strip()
        ]
        if criteria_rows:
            parsed_criteria = []
            for row in criteria_rows:
                name = str(row[0]).strip()
                if not name or name.lower() in {"criterio", "criterios", "total", "#", "datos"}:
                    continue
                values = []
                for v in row[1:]:
                    if v is None:
                        continue
                    try:
                        values.append(float(v))
                    except Exception:
                        continue
                # accept only if values likely represent per-material array
                if values and ((material_names and len(values) == len(material_names)) or (not material_names and len(values) > 1)):
                    parsed_criteria.append((name, values))

            if parsed_criteria:
                criteria = [name for name, _ in parsed_criteria]
                for name, values in parsed_criteria:
                    matrix[name] = values

# fallback material names
if not material_names:
    preferred = [s for s in wb.worksheets if ("matriz" in s.title.lower() and "datos" in s.title.lower())]
    if not preferred:
        preferred = [s for s in wb.worksheets if any(k in s.title.lower() for k in ("matriz", "datos"))]
    others = [s for s in wb.worksheets if s not in preferred]
    search_order = preferred + others
    for sheet in search_order:
        # debug: show which sheet we're checking for material header
        # print('checking sheet for materials:', sheet.title)
        header_row_index = None
        header = None
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
            if any(isinstance(c, str) and 'material' in c.lower() for c in row if c is not None):
                header_row_index = i
                header = row
                break
        if header_row_index is None:
            continue

        try:
            material_col_index = next(idx for idx, c in enumerate(header) if isinstance(c, str) and 'material' in c.lower())
        except StopIteration:
            continue

        names = []
        for row in sheet.iter_rows(min_row=header_row_index + 1, max_row=sheet.max_row, values_only=True):
            v = row[material_col_index] if material_col_index < len(row) else None
            if v is None:
                continue
            s = str(v).strip()
            if s and s.lower() not in {"material", "materiales", "total", "#", "datos"}:
                names.append(s)
        if names:
            material_names = names
            break

if not material_names:
    raise SystemExit('No materials found')

# row-based criteria parsing
if not criteria:
    for sheet in wb.worksheets:
        parsed = [
            row
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True)
            if row and row[0] is not None and str(row[0]).strip()
        ]
        parsed_criteria = []
        for row in parsed:
            name = str(row[0]).strip()
            if not name or name.lower() in {"criterio", "criterios", "total", "#", "datos"}:
                continue
            values = []
            for v in row[1:]:
                if v is None:
                    continue
                try:
                    values.append(float(v))
                except Exception:
                    continue
            if values:
                parsed_criteria.append((name, values))

        if parsed_criteria:
            criteria = [name for name, _ in parsed_criteria]
            for name, values in parsed_criteria:
                matrix[name] = values
            break

# column-based criteria parsing
if not criteria and material_names:
    preferred = [s for s in wb.worksheets if ("matriz" in s.title.lower() and "datos" in s.title.lower())]
    if not preferred:
        preferred = [s for s in wb.worksheets if any(k in s.title.lower() for k in ("matriz", "datos"))]
    others = [s for s in wb.worksheets if s not in preferred]
    search_order = preferred + others
    for sheet in search_order:
        header_row_index = None
        header = None
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
            if any(isinstance(c, str) and 'material' in c.lower() for c in row if c is not None):
                header_row_index = i
                header = row
                break

        if header_row_index is None:
            continue

        try:
            material_col_index = next(idx for idx, c in enumerate(header) if isinstance(c, str) and 'material' in c.lower())
        except StopIteration:
            continue

        found = False
        for col_idx in range(material_col_index + 1, len(header)):
            name = header[col_idx]
            if name is None:
                continue
            name_str = str(name).strip()
            if not name_str:
                continue
            values = []
            for row in sheet.iter_rows(min_row=header_row_index + 1, max_row=sheet.max_row, values_only=True):
                v = row[col_idx] if col_idx < len(row) else None
                if v is None:
                    continue
                try:
                    values.append(float(v))
                except Exception:
                    continue
            # accept columns only when they provide numeric arrays matching materials
            if values and len(values) == len(material_names):
                matrix[name_str] = values
                criteria.append(name_str)
                found = True

        if found:
            break

print('materials:', material_names)
print('criteria:', criteria)
for k, v in matrix.items():
    print(k, v[:6])
